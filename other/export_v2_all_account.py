# -*- coding: utf-8 -*-
#
import base64
import datetime
import uuid

import openpyxl
import pymysql

from collections import OrderedDict

from Cryptodome.Cipher import AES
from Cryptodome.Util.Padding import pad
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


class ToolException(Exception):
    pass


class JMSDecryptHandler:
    def __init__(self, secret_key):
        self.key = secret_key

    @property
    def real_key(self):
        key = self.key if isinstance(self.key, bytes) else bytes(self.key, encoding='utf-8')
        return key[:32] if len(key) >= 32 else pad(key, 32)

    def decrypt(self, text):
        if not text:
            return text
        try:
            metadata = text[:72]
            header = base64.b64decode(metadata[:24])
            nonce = base64.b64decode(metadata[24:48])
            tag = base64.b64decode(metadata[48:])
            ciphertext = base64.b64decode(text[72:])
            cipher = AES.new(self.real_key, AES.MODE_GCM, nonce=nonce)
            cipher.update(header)
            plain_text_bytes = cipher.decrypt_and_verify(ciphertext, tag)
        except Exception:
            # 解密失败
            plain_text_bytes = b'\xe8\xa7\xa3\xe5\xaf\x86\xe5\xa4\xb1\xe8\xb4\xa5'
        return plain_text_bytes.decode('utf-8')


class Tool(object):
    def __init__(self, db_info: dict, secret_key: str):
        self.parser = JMSDecryptHandler(secret_key=secret_key)
        self.__db_info = db_info
        self.__conn = None
        self.__cursor = None
        self.__data = []
        self.empty = '无'

    def __del__(self):
        try:
            self.cursor.close()
            self.conn.close()
        except Exception:
            pass

    @property
    def conn(self):
        if self.__conn is None:
            self.__conn = self.__get_mysql_conn(**self.__db_info)
        return self.__conn

    @property
    def cursor(self):
        if self.__cursor is None:
            self.__cursor = self.conn.cursor()
        return self.__cursor

    @staticmethod
    def __get_mysql_conn(**kwargs):
        try:
            conn = pymysql.connect(
                **kwargs, cursorclass=pymysql.cursors.DictCursor
            )
        except pymysql.MySQLError as error:
            raise ToolException(f'连接数据库失败: {error}')
        return conn

    def __execute_sql(self, sql, args=tuple(), many=True):
        self.cursor.execute(sql, args)
        return self.cursor.fetchall() if many else self.cursor.fetchone()

    @property
    def header(self):
        """
        获取主机名、IP、系统用户名称、账号用户名、明文密码/密钥，该账号最近一次成功登录该资产的时间
        """
        return OrderedDict({
            'id': '主机ID', 'hostname': '主机名', 'ip': 'IP', 'asset_protocol': '资产协议',
            'system_user': '系统用户名称', 'system_user_protocol': '系统用户协议',
            'username': '账号用户名', 'password': '密码', 'private_key': '秘钥', 'public_key': '公钥',
            'login_date': '最近一次登录时间'
        })

    @staticmethod
    def __auto_adjust_column_width(ws):
        def calc_handler(x):
            return len(x.encode()) if x else 0

        pre_rows = ws.iter_cols(min_row=1, max_row=2, values_only=True)
        # 动态分配表头宽度
        column_width = [max(map(calc_handler, i)) for i in pre_rows]
        # 调整首行宽度
        for i, width in enumerate(column_width, 1):
            width = width if width < 100 else 100
            width = 10 if width < 10 else width
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = width

    @staticmethod
    def __fill_row_color(ws, row_numbers):
        fill = PatternFill(start_color='cccccc', end_color='cccccc', fill_type='solid')
        for n in row_numbers:
            for row in ws.iter_rows(min_row=n, max_row=n):
                for cell in row:
                    cell.fill = fill

    @staticmethod
    def uuid_str(u):
        return str(uuid.UUID(u))

    def __build_excel_data(self, row_obj):
        if isinstance(row_obj, dict):
            row_obj = [row_obj[h] or self.empty for h in self.header.keys()]
        self.__data.append(row_obj)

    def __get_assets(self):
        sql = 'SELECT id, ip, hostname, protocols FROM assets_asset ORDER BY hostname, ip;'
        return self.__execute_sql(sql)

    def __get_system_user(self, asset_id):
        sql = 'SELECT a.name AS a_name, a.username AS a_username, a.password AS a_password, ' \
              'a.private_key AS a_private_key, a.public_key AS a_public_key, ' \
              'b.name AS b_name, b.username AS b_username, b.password AS b_password, ' \
              'b.private_key AS b_private_key, b.public_key AS b_public_key, b.protocol AS system_user_protocol ' \
              'FROM assets_authbook a INNER JOIN assets_systemuser b ON a.systemuser_id = b.id ' \
              'WHERE a.asset_id=%s;'
        return self.__execute_sql(sql, (asset_id,))

    def __get_asset_last_login_time(self, asset_id):
        sql = 'SELECT date_start FROM terminal_session WHERE asset_id=%s ORDER BY date_start LIMIT 1;'
        resp = self.__execute_sql(sql, (self.uuid_str(asset_id),), many=False) or {}
        return str(resp.get('date_start')) if resp.get('date_start') else None

    @staticmethod
    def __guess_value(obj, name, prefix=('a', 'b')):
        value = None
        for p in prefix:
            if value := obj.get(f'{p}_{name}'):
                break
        return value

    def get_data(self):
        for asset in self.__get_assets():
            print(f"获取资产 [{asset['hostname']}({asset['ip']})] 的信息")
            last_login = self.__get_asset_last_login_time(asset['id'])
            row_obj = {
                'id': asset['id'], 'hostname': asset['hostname'], 'ip': asset['ip'],
                'asset_protocol': asset['protocols'], 'login_date': last_login or self.empty
            }
            system_users = self.__get_system_user(asset['id'])
            if not system_users:
                system_users = [{
                    'a_id': None, 'a_name': None, 'a_username': None, 'a_password': None,
                    'system_user_protocol': None, 'a_private_key': None, 'a_public_key': None
                }]
            for su in system_users:
                row_obj['system_user'] = self.__guess_value(su, 'name')
                row_obj['username'] = self.__guess_value(su, 'username')
                row_obj['system_user_protocol'] = su['system_user_protocol']
                row_obj['password'] = self.parser.decrypt(self.__guess_value(su, 'password'))
                row_obj['private_key'] = self.parser.decrypt(self.__guess_value(su, 'private_key'))
                row_obj['public_key'] = self.parser.decrypt(self.__guess_value(su, 'public_key'))
                self.__build_excel_data(row_obj)

            # 一个资产可能会有多个账号，如果想要让每个资产中间加空格，解开下方代码注释即可
            # self.__build_excel_data([])

    def save(self):
        date_string = datetime.datetime.now().strftime('%Y-%m-%d')
        filename = f'./JumpServer-Asset-{date_string}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.freeze_panes = 'A2'
        ws.append(list(self.header.values()))
        need_fill_row = []
        for i, d in enumerate(self.__data, 2):
            if len(d) < 1:
                need_fill_row.append(i)
            ws.append(d)
        self.__auto_adjust_column_width(ws)
        # 可以给指定的行填充颜色
        # self.__fill_row_color(ws, need_fill_row)
        wb.save(filename)

    def run(self):
        try:
            self.get_data()
        except ToolException as error:
            print(error)
            return
        else:
            self.save()


t = Tool(
    db_info={
        'host': 'mysql_host', 'user': 'mysql_username',
        'password': 'mysql_password', 'db': 'mysql_db_name'
    },
    # 用来解密数据库中加密数据的
    secret_key='Jumpserver_Config_SECRET_KEY'
)
t.run()
